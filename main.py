"""
main.py - Render Video Reup Pro
Flow: CSV → download temp → FFmpeg → lưu (upload Drive | move vào thư mục local)
      → output CSV (xoá cột video_url, thay bằng cột 'Link Video' = link Drive/local path, lỗi = "Lỗi", chưa chạy = rỗng)
"""

from __future__ import annotations   # cho phép 'X | None' chạy trên Python 3.8/3.9

import csv
import io
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from pathlib import Path

# ── Fix X11 BadLength (RenderAddGlyphs) — Layer 1: env vars trước khi import Tk ───────
# Ép Xft dùng 1-bit monochrome glyphs (giảm 97% kích thước bitmap) & 24-bit visuals
if sys.platform != "win32":
    os.environ["XFT_ANTIALIAS"] = "0"
    os.environ["XFT_HINTING"] = "0"
    os.environ["XFT_RGBA"] = "none"
    os.environ["XFT_MAX_GLYPH_MEMORY"] = "10485760"
    os.environ["XLIB_SKIP_ARGB_VISUALS"] = "1"
    os.environ["GDK_SCALE"] = "1"
    os.environ["GDK_DPI_SCALE"] = "1"
    os.environ.pop("WAYLAND_DISPLAY", None)
# ────────────────────────────────────────────────────────────────────────────

from tkinter import BooleanVar, StringVar, filedialog, messagebox

import customtkinter as ctk
import requests
from PIL import Image as PILImage

import license as license_mod
import single_instance

# Vô hiệu hóa load_font TTF ngoài của CTk trên Linux để tránh bão font handles Xft
if sys.platform != "win32":
    try:
        import customtkinter.windows.widgets.utility.font_manager as ctk_fm
        ctk_fm.FontManager.load_font = lambda *args, **kwargs: True
    except Exception:
        pass

APP_VERSION = "v1.0.0"
IS_PRO = (getattr(license_mod, "APP_ID", "") == "tool_reup_video_pro")
APP_TITLE = (
    f"Render Video Reup Pro {APP_VERSION}"
    if IS_PRO
    else f"Render Video Reup {APP_VERSION}"
)

# ── Fix X11 BadLength — Layer 2: tắt CTk DPI auto-detect trước khi tạo bất kỳ cửa sổ nào ──
# CTk 5.2+ có _check_dpi_scaling chạy định kỳ và override tk.scaling → gây crash.
if sys.platform != "win32":
    try:
        ctk.deactivate_automatic_dpi_awareness()   # CTk >= 5.2.0
    except AttributeError:
        pass
    try:
        ctk.set_widget_scaling(1.0)
        ctk.set_window_scaling(1.0)
    except Exception:
        pass
# ─────────────────────────────────────────────────────────────────────────────

from core import FFMPEG, VideoDownloader, VideoProcessor
from tts import (DEFAULT_PROMPT, DEFAULT_AUTOVOICE_VOICE, DEFAULT_VIDEOAI_VOICE,
                 DEFAULT_VOICE_LABEL, VOICE_CHOICES, make_voice)

# ── Theme ────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

_WIN_FLAGS = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0


def _apply_linux_x11_fix(tk_root) -> None:
    """Áp dụng workaround triệt để cho lỗi X11 BadLength - RenderAddGlyphs trên Linux.

    Fix bao gồm các bước khi mỗi cửa sổ CTk được tạo:
    1. Tắt CTk DPI auto-detect & ép widget/window scaling = 1.0.
    2. Ép Tk internal scaling = 1.0 → glyph nhỏ hơn → request X11 không vượt giới hạn (64KB).
    3. Tự động phát hiện font hệ thống có sẵn (Ubuntu, Liberation Sans, FreeSans, Arial) → fallback an toàn.
    """
    if sys.platform == "win32":
        return
    try:
        try:
            ctk.deactivate_automatic_dpi_awareness()
        except AttributeError:
            pass
        try:
            ctk.set_widget_scaling(1.0)
            ctk.set_window_scaling(1.0)
        except Exception:
            pass
        try:
            tk_root.tk.call("tk", "scaling", 1.0)
            tk_root.tk.call("tk", "scaling", "-displayof", ".", 1.0)
        except Exception:
            pass
        import tkinter.font as tkfont
        try:
            avail = set(tkfont.families(tk_root))
            chosen = "sans-serif"
            for f_cand in ("Ubuntu", "Liberation Sans", "DejaVu Sans", "FreeSans", "Arial", "sans-serif"):
                if f_cand in avail:
                    chosen = f_cand
                    break
            for fn in ("TkDefaultFont", "TkTextFont", "TkFixedFont", "TkMenuFont", "TkHeadingFont", "TkCaptionFont", "TkTooltipFont"):
                try:
                    tkfont.nametofont(fn).configure(family=chosen, size=10)
                except Exception:
                    pass
            try:
                ctk.FontManager._font_family = chosen
            except Exception:
                pass
        except Exception:
            pass
        # Vô hiệu hoá hoàn toàn vòng lặp kiểm tra DPI định kỳ của CTk
        # (belt-and-suspenders thêm vào deactivate_automatic_dpi_awareness).
        # Nếu _check_dpi_scaling chạy lại sẽ override tk.scaling=1.0 → X11 BadLength crash.
        try:
            ctk.CTk._check_dpi_scaling = lambda self: None
        except Exception:
            pass
    except Exception:
        pass


# ── Cấu hình kiểm tra license định kỳ (ép thời gian ngắn lại khi test) ────────
_RECHECK_SECONDS = int(os.environ.get("LICENSE_RECHECK_SECONDS", 30 * 60))  # 30'
_RETRY_SECONDS   = int(os.environ.get("LICENSE_RETRY_SECONDS",   60))       # 1'
_MAX_RETRIES     = int(os.environ.get("LICENSE_MAX_RETRIES",     3))        # 3 lần


def _resource(name: str) -> Path:
    """Locate a bundled resource (works in dev and inside a PyInstaller exe)."""
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
    return base / name


def _crash_dir() -> Path:
    """Thư mục GHI ĐƯỢC để lưu crash.log (AppImage-aware qua license_mod)."""
    try:
        return license_mod._exe_dir()
    except Exception:
        return Path(__file__).parent


def _write_crash(tb: str) -> None:
    """Ghi traceback ra crash.log để chẩn đoán khi chạy không có console."""
    try:
        (_crash_dir() / "crash.log").write_text(tb, encoding="utf-8")
    except Exception:
        pass

STATUS_ICON = {
    "pending":     "...",
    "downloading": "DL",
    "processing":  "*",
    "uploading":   "@",
    "done":        "OK",
    "warning":     "!",
    "error":       "X",
}


# ── Phân loại log ─────────────────────────────────────────────────────────────
# Mỗi dòng log gắn 1 nhãn để lọc bằng dropdown. Dòng hệ thống (cat=None) chỉ
# xuất hiện ở mục "Tất cả".
LOG_DOWNLOAD = "Tải video"
LOG_TEXT     = "Tạo text"
LOG_VOICE    = "Tạo voice"
LOG_PROCESS  = "Xử lý video"
LOG_UPLOAD   = "Upload"
LOG_RESULT   = "Kết quả"
LOG_ERROR    = "Lỗi"
LOG_CATS        = (LOG_DOWNLOAD, LOG_TEXT, LOG_VOICE, LOG_PROCESS, LOG_UPLOAD, LOG_RESULT, LOG_ERROR)
LOG_FILTER_ALL  = "Tất cả"
LOG_FILTER_OPTS = (LOG_FILTER_ALL,) + LOG_CATS


# ── Hiển thị danh sách (phân trang) ───────────────────────────────────────────
# Chỉ render mỗi trang _PAGE_SIZE dòng để tránh treo khi CSV có hàng nghìn video.
_PAGE_SIZE = 20

# ── Thumbnail helpers ─────────────────────────────────────────────────────────

_THUMB_SIZE  = (64, 64)
# Giới hạn số ảnh tải đồng thời (tránh bão thread/mạng khi mở 1 trang nhiều ảnh).
_THUMB_SEM   = threading.Semaphore(8)
_THUMB_HEADS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Referer":    "https://shopee.vn/",
}
_PLACEHOLDER = ctk.CTkImage(
    light_image=PILImage.new("RGB", _THUMB_SIZE, (60, 60, 60)),
    dark_image =PILImage.new("RGB", _THUMB_SIZE, (60, 60, 60)),
    size=_THUMB_SIZE,
)


# ── VideoRow ─────────────────────────────────────────────────────────────────

class VideoRow(ctk.CTkFrame):
    """One row: checkbox | status | thumbnail 64×64 | product name  (80px)"""

    _ROW_H = 80

    def __init__(self, parent, app, video_data: dict, global_index: int, **kw):
        bg = "#2a2a2a" if global_index % 2 == 0 else "#313131"
        super().__init__(parent, fg_color=bg, corner_radius=4,
                         height=self._ROW_H, **kw)
        self.pack_propagate(False)

        self._app         = app
        self.global_index = global_index
        self.video_data   = video_data
        # Lựa chọn lấy từ model (giữ nguyên khi chuyển trang qua lại)
        self.selected     = BooleanVar(value=app._selected[global_index])
        self._ctk_img     = None
        self._alive       = True   # Guard: set False in destroy() to stop pending async thumb callbacks

        self.chk = ctk.CTkCheckBox(
            self, text="", variable=self.selected, command=self._on_toggle,
            width=28, checkbox_width=18, checkbox_height=18,
        )
        self.chk.pack(side="left", padx=(8, 2))

        self.lbl_status = ctk.CTkLabel(self, text="...", width=26, font=("", 14))
        self.lbl_status.pack(side="left", padx=2)

        self.thumb = ctk.CTkLabel(
            self, image=_PLACEHOLDER, text="",
            width=_THUMB_SIZE[0], height=_THUMB_SIZE[1],
        )
        self.thumb.pack(side="left", padx=(6, 8), pady=8)

        val = video_data.get("product_name") if video_data.get("product_name") is not None else video_data.get("productName")
        p_name = "" if val is None or str(val).strip().lower() in ("", "nan", "none") else str(val).strip()
        name = p_name if p_name else "(dữ liệu product_name rỗng)"
        # TRUNCATE to avoid X11 BadLength (RenderAddGlyphs) crash on very long CSV strings
        if len(name) > 200:
            name = name[:197] + "..."
        self.name_lbl = ctk.CTkLabel(
            self, text=name, anchor="w",
            font=("", 12), wraplength=500, justify="left",
        )
        self.name_lbl.pack(side="left", fill="x", expand=True, padx=(0, 8))

        self.pb = ctk.CTkProgressBar(self, width=110, height=8)
        self.pb.set(0)

        # Khôi phục trạng thái đã lưu (khi quay lại trang đã xử lý xong)
        self.set_status(app._status[global_index])

        url = video_data.get("link_image", "")
        if url:
            threading.Thread(target=self._load_thumb, args=(url,),
                             daemon=True).start()

    def destroy(self):
        """Đánh dấu widget là đã bị huỷ trước khi teardown — ngăn các callback async thumbnail."""
        self._alive = False
        super().destroy()

    def _on_toggle(self):
        # Ghi lựa chọn trở lại model để giữ qua các trang
        self._app._selected[self.global_index] = self.selected.get()

    def _load_thumb(self, url: str):
        with _THUMB_SEM:                     # giới hạn số ảnh tải đồng thời
            try:
                resp = requests.get(url, headers=_THUMB_HEADS, timeout=10)
                resp.raise_for_status()
                img = PILImage.open(io.BytesIO(resp.content)).convert("RGB")
                img = img.resize(_THUMB_SIZE, PILImage.LANCZOS)
                ctk_img = ctk.CTkImage(light_image=img, dark_image=img,
                                       size=_THUMB_SIZE)
                if not self._alive:          # widget đã bị destroy → bỏ qua
                    return
                try:
                    self.after(0, lambda i=ctk_img: self._apply_thumb(i))
                except Exception:            # TclError khi widget bị destroy giữa chừng
                    pass
            except Exception:
                pass

    def _apply_thumb(self, ctk_img: ctk.CTkImage):
        if not self._alive:                  # widget đã bị destroy trước khi callback chạy
            return
        try:
            self._ctk_img = ctk_img
            self.thumb.configure(image=ctk_img)
        except Exception:
            pass

    def set_status(self, status: str):
        self.lbl_status.configure(text=STATUS_ICON.get(status, "?"))
        if status in ("downloading", "processing", "uploading"):
            self.pb.pack(side="right", padx=8)
        else:
            try:
                self.pb.pack_forget()
            except Exception:
                pass

    def set_progress(self, value: float):
        self.pb.set(max(0.0, min(1.0, value)))


# ── Main App ─────────────────────────────────────────────────────────────────

class App(ctk.CTk):

    def __init__(self, license_data: dict | None = None):
        super().__init__()
        _apply_linux_x11_fix(self)
        # Belt-and-suspenders: vô hiệu hoá vòng kiểm tra DPI định kỳ ở instance level
        # (bổ sung cho class-level patch trong _apply_linux_x11_fix)
        if sys.platform != "win32":
            try:
                self._check_dpi_scaling = lambda: None
            except Exception:
                pass
        self._license_data = license_data or {}
        self.title(APP_TITLE)
        self.geometry("1220x820")
        self.minsize(960, 640)
        self._set_app_icon()

        self._videos:    list[dict]     = []
        self._rows:      list[VideoRow] = []     # các dòng đang hiển thị (1 trang)
        # Trạng thái theo từng video — giữ độc lập với widget để phân trang được
        self._selected:  list[bool]     = []     # tick chọn / không
        self._status:    list[str]      = []     # "pending"/"done"/...
        self._row_by_index: dict[int, VideoRow] = {}  # global index → widget hiện thị
        self._page       = 0                     # trang đang xem (0-based)
        self._csv_path:  str            = ""
        self._stop_flag  = False
        self._processing = False

        self._downloader = VideoDownloader()
        self._processor  = VideoProcessor()

        self._creds_path   = StringVar(value="")
        self._audio_path   = StringVar(value="")
        self._logo_path    = StringVar(value="")
        self._csv_out_path = StringVar(value="")
        self._local_dir    = StringVar(value="")

        self._login_in_progress = False
        self._login_cancel_flag = threading.Event()

        self._build_ui()
        self._check_ffmpeg()

        # ── Trạng thái kiểm tra license định kỳ ──
        self._recheck_job      = None    # id job after() để có thể hủy
        self._recheck_retries  = 0       # đếm số lần retry khi mất mạng
        self._license_locked   = False   # tránh khóa UI nhiều lần
        self._relaunch_license = False   # cờ báo main() mở lại LicenseDialog

        # Hẹn lần kiểm tra đầu tiên (mặc định sau 30 phút)
        self._recheck_job = self.after(_RECHECK_SECONDS * 1000,
                                       self._schedule_recheck)

    # ── App icon ─────────────────────────────────────────────────────────────

    def _set_app_icon(self):
        ico = _resource("video.ico")
        if not ico.exists():
            return
        try:
            self.iconbitmap(str(ico))
            # CTk reapplies its own icon shortly after init — re-set then.
            self.after(250, lambda: self.iconbitmap(str(ico)))
        except Exception:
            pass

    # ── FFmpeg check ─────────────────────────────────────────────────────────

    def _check_ffmpeg(self):
        try:
            subprocess.run([FFMPEG, "-version"], capture_output=True,
                           creationflags=_WIN_FLAGS, check=True)
        except (FileNotFoundError, subprocess.CalledProcessError):
            messagebox.showerror(
                "FFmpeg chưa cài",
                "Không tìm thấy ffmpeg!\n\nChạy lại run.bat để tự động tải,\n"
                "hoặc chạy thủ công:\n  python setup_ffmpeg.py",
            )

    # ── Bắt lỗi trong callback GUI (nút bấm, sự kiện) ──────────────────────────
    def report_callback_exception(self, exc, val, tb_obj):
        """Lỗi Python trong callback → ghi crash.log + hiện hộp thoại, KHÔNG đóng app."""
        import traceback
        tb = "".join(traceback.format_exception(exc, val, tb_obj))
        _write_crash(tb)
        sys.stderr.write(tb)
        try:
            messagebox.showerror(
                "Lỗi", f"{getattr(exc, '__name__', exc)}: {val}\n\n"
                       "Chi tiết đã ghi vào crash.log (cạnh file .AppImage).")
        except Exception:
            pass

    # ── UI ───────────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self._build_topbar()
        self._build_settings()
        self._build_video_list()
        self._build_bottom()

    # ── Top bar ──────────────────────────────────────────────────────────────

    def _build_topbar(self):
        bar = ctk.CTkFrame(self, height=52, corner_radius=0, fg_color="#1a1a2e")
        bar.grid(row=0, column=0, columnspan=2, sticky="ew")
        bar.grid_propagate(False)

        ctk.CTkLabel(bar, text=f" {APP_TITLE}",
                     font=("", 18, "bold")).pack(side="left", padx=16)

        self._csv_label = ctk.CTkLabel(bar, text="Chưa import CSV/Excel",
                                        text_color="gray", font=("", 12))
        self._csv_label.pack(side="left", padx=6)

        ctk.CTkButton(bar, text=" Import CSV/Excel", width=130, height=32,
                      command=self._import_file).pack(side="left", padx=6)

        # CSV output path
        ctk.CTkButton(bar, text=" Lưu CSV", width=110, height=32,
                      command=self._pick_csv_out).pack(side="right", padx=10)

        self._csv_out_label = ctk.CTkLabel(bar, text="Chưa chọn nơi lưu CSV",
                                            text_color="gray", font=("", 12))
        self._csv_out_label.pack(side="right", padx=6)

    # ── Settings panel ───────────────────────────────────────────────────────

    def _build_settings(self):
        pane = ctk.CTkScrollableFrame(self, width=250,
                                       label_text="*  Cài đặt xử lý")
        pane.grid(row=1, column=0, sticky="nsew", padx=(8, 4), pady=8)

        # ── Concurrency (Chỉ có ở bản Pro) ──
        if IS_PRO:
            self._section(pane, "  Xử lý đồng thời")
            self._workers = self._labeled_entry(pane, "Số video cùng lúc:", "2")
            self._delay   = self._labeled_entry(pane, "Giãn cách mỗi video (giây):", "1")
        else:
            self._workers = None
            self._delay   = None

        # ── Trim ──
        self._section(pane, "  Cắt video")
        self._trim_start = self._labeled_entry(pane, "Cắt đầu (giây):", "0")
        self._trim_end   = self._labeled_entry(pane, "Cắt đuôi (giây):", "0")

        # ── Audio ──
        self._section(pane, "  Thay Voice")

        self._audio_label = ctk.CTkLabel(pane, text="Chưa chọn file audio",
                                          text_color="gray", font=("", 11),
                                          wraplength=230, justify="left")
        self._audio_label.pack(anchor="w", pady=(0, 4))

        row = ctk.CTkFrame(pane, fg_color="transparent")
        row.pack(fill="x", pady=2)
        self._btn_pick_audio = ctk.CTkButton(
            row, text=" Chọn Audio", height=30, command=self._pick_audio)
        self._btn_pick_audio.pack(side="left", fill="x", expand=True,
                                  padx=(0, 2))
        self._btn_clear_audio = ctk.CTkButton(
            row, text="X", width=30, height=30, fg_color="#555",
            command=self._clear_audio)
        self._btn_clear_audio.pack(side="right")

        self._mute_var = BooleanVar(value=False)
        self._mute_chk = ctk.CTkCheckBox(pane, text="Tắt tiếng (mute)",
                                         variable=self._mute_var)
        self._mute_chk.pack(anchor="w", pady=(4, 0))

        # ── Voice AI (prompt -> Gemini -> Google TTS) ──
        self._voiceai_var = BooleanVar(value=False)
        ctk.CTkSwitch(
            pane, text=" Voice AI (Gemini + Google TTS)",
            variable=self._voiceai_var, command=self._toggle_voiceai,
        ).pack(anchor="w", pady=(10, 2))

        self._voiceai_box = ctk.CTkFrame(pane, fg_color="transparent")
        # Ẩn ban đầu; hiện khi bật switch

        # AI tạo text: Gemini hoặc ChatGPT
        arow = ctk.CTkFrame(self._voiceai_box, fg_color="transparent")
        arow.pack(fill="x", pady=(6, 2))
        ctk.CTkLabel(arow, text="AI tạo text:", width=85,
                     anchor="w").pack(side="left")
        self._ai_provider = ctk.CTkOptionMenu(
            arow, values=["Gemini", "ChatGPT"],
            width=145, command=self._toggle_ai_provider)
        self._ai_provider.set("Gemini")
        self._ai_provider.pack(side="right")

        # Khối Gemini
        self._gemini_box = ctk.CTkFrame(self._voiceai_box, fg_color="transparent")
        self._gemini_key = self._fullwidth_entry(
            self._gemini_box, "API Key Gemini:",
            placeholder="Dán API key Gemini")
        self._gemini_model = self._fullwidth_entry(
            self._gemini_box, "Model Gemini:",
            placeholder="gemini-3.1-flash-lite", default="gemini-3.1-flash-lite")

        # Khối ChatGPT
        self._chatgpt_box = ctk.CTkFrame(self._voiceai_box, fg_color="transparent")
        self._chatgpt_key = self._fullwidth_entry(
            self._chatgpt_box, "API Key OpenAI:",
            placeholder="sk-...")
        self._chatgpt_model = self._fullwidth_entry(
            self._chatgpt_box, "Model ChatGPT:",
            placeholder="gpt-4o-mini", default="gpt-4o-mini")

        self._ai_provider_anchor = ctk.CTkFrame(self._voiceai_box, height=0,
                                                fg_color="transparent")
        self._ai_provider_anchor.pack(fill="x")

        # Nhà cung cấp giọng
        prow = ctk.CTkFrame(self._voiceai_box, fg_color="transparent")
        prow.pack(fill="x", pady=(6, 2))
        ctk.CTkLabel(prow, text="AI tạo voice:", width=85,
                     anchor="w").pack(side="left")
        self._tts_provider = ctk.CTkOptionMenu(
            prow, values=["Google TTS", "AutoVoice.vn", "Voice API (videoai)"],
            width=145, command=self._toggle_tts_provider)
        self._tts_provider.set("Google TTS")
        self._tts_provider.pack(side="right")

        # ── Khối Google TTS ──
        self._google_box = ctk.CTkFrame(self._voiceai_box, fg_color="transparent")
        self._tts_key = self._fullwidth_entry(
            self._google_box, "API Key Google TTS:",
            placeholder="Dán API key Google TTS")
        vrow = ctk.CTkFrame(self._google_box, fg_color="transparent")
        vrow.pack(fill="x", pady=(6, 2))
        ctk.CTkLabel(vrow, text="Giọng:", width=55, anchor="w").pack(side="left")
        self._voice_name = ctk.CTkOptionMenu(
            vrow, values=list(VOICE_CHOICES.keys()), width=175)
        self._voice_name.set(DEFAULT_VOICE_LABEL)
        self._voice_name.pack(side="right")
        self._tts_speed = self._labeled_entry(
            self._google_box, "Tốc độ đọc:", "1.2")

        # ── Khối AutoVoice.vn ──
        self._autovoice_box = ctk.CTkFrame(self._voiceai_box, fg_color="transparent")
        self._autovoice_key = self._fullwidth_entry(
            self._autovoice_box, "X-API-Key:",
            placeholder="Dán API key AutoVoice")
        self._autovoice_voice = self._fullwidth_entry(
            self._autovoice_box, "Giọng (voiceId):",
            placeholder=DEFAULT_AUTOVOICE_VOICE, default=DEFAULT_AUTOVOICE_VOICE)
        self._autovoice_speed = self._labeled_entry(
            self._autovoice_box, "Tốc độ đọc:", "1")

        # ── Khối Voice API (videoai) ──
        self._videoai_box = ctk.CTkFrame(self._voiceai_box, fg_color="transparent")
        self._videoai_key = self._fullwidth_entry(
            self._videoai_box, "X-API-Key:",
            placeholder="Dán API key Voice API")
        self._videoai_voice = self._fullwidth_entry(
            self._videoai_box, "Giọng (voice_name):",
            placeholder=DEFAULT_VIDEOAI_VOICE, default=DEFAULT_VIDEOAI_VOICE)
        self._videoai_speed = self._labeled_entry(
            self._videoai_box, "Tốc độ đọc:", "1")

        # Mốc để pack khối nhà cung cấp đúng chỗ (ngay trên Prompt)
        self._tts_provider_anchor = ctk.CTkFrame(self._voiceai_box, height=0,
                                                 fg_color="transparent")
        self._tts_provider_anchor.pack(fill="x")

        ctk.CTkLabel(self._voiceai_box, text="Prompt tạo lời thoại:",
                     anchor="w").pack(anchor="w", pady=(6, 2))
        ctk.CTkLabel(
            self._voiceai_box,
            text="Chèn dữ liệu CSV bằng ${tên_cột} (vd ${product_name}).",
            text_color="gray", font=("", 10), wraplength=230, justify="left",
        ).pack(anchor="w")
        self._prompt_box = ctk.CTkTextbox(self._voiceai_box, height=120,
                                          wrap="word")
        self._prompt_box.pack(fill="x", pady=(2, 0))
        self._prompt_box.insert("1.0", DEFAULT_PROMPT)

        # Mốc vị trí để pack khối Voice AI đúng chỗ khi bật/tắt switch
        self._voiceai_anchor = ctk.CTkFrame(pane, height=0,
                                            fg_color="transparent")
        self._voiceai_anchor.pack(fill="x")
        self._toggle_ai_provider()  # hiện đúng khối AI tạo text ban đầu
        self._toggle_tts_provider()  # hiện đúng khối nhà cung cấp ban đầu
        self._toggle_voiceai()  # áp dụng trạng thái ẩn ban đầu

        # ── Logo ──
        self._section(pane, "  Logo / Watermark")

        self._logo_label = ctk.CTkLabel(pane, text="Chưa chọn file logo",
                                         text_color="gray", font=("", 11),
                                         wraplength=230, justify="left")
        self._logo_label.pack(anchor="w", pady=(0, 4))

        row2 = ctk.CTkFrame(pane, fg_color="transparent")
        row2.pack(fill="x", pady=2)
        ctk.CTkButton(row2, text=" Chọn Logo", height=30,
                      command=self._pick_logo).pack(side="left", fill="x",
                                                     expand=True, padx=(0, 2))
        ctk.CTkButton(row2, text="X", width=30, height=30, fg_color="#555",
                      command=self._clear_logo).pack(side="right")

        pos_row = ctk.CTkFrame(pane, fg_color="transparent")
        pos_row.pack(fill="x", pady=(8, 2))
        ctk.CTkLabel(pos_row, text="Vị trí:", width=90,
                     anchor="w").pack(side="left")
        self._logo_pos = ctk.CTkOptionMenu(
            pos_row,
            values=["Top-Right", "Top-Left", "Bottom-Right", "Bottom-Left", "Center"],
            width=165,
        )
        self._logo_pos.pack(side="right")

        # Chuyển động: cố định / chạy vòng theo viền / nảy DVD.
        # Vị trí ở trên đóng vai trò điểm xuất phát khi có chuyển động.
        motion_row = ctk.CTkFrame(pane, fg_color="transparent")
        motion_row.pack(fill="x", pady=2)
        ctk.CTkLabel(motion_row, text="Chuyển động:", width=90,
                     anchor="w").pack(side="left")
        self._logo_motion = ctk.CTkOptionMenu(
            motion_row,
            values=["Cố định", "Chạy vòng theo viền", "Nảy DVD"],
            width=165,
            command=self._toggle_logo_speed,
        )
        self._logo_motion.pack(side="right")

        # Tốc độ — chỉ hiện khi Chuyển động khác "Cố định"
        self._logo_speed_row = ctk.CTkFrame(pane, fg_color="transparent")
        ctk.CTkLabel(self._logo_speed_row, text="Tốc độ:", width=90,
                     anchor="w").pack(side="left")
        self._logo_speed = ctk.CTkOptionMenu(
            self._logo_speed_row, values=["Chậm", "Vừa", "Nhanh"], width=165)
        self._logo_speed.set("Vừa")
        self._logo_speed.pack(side="right")
        # Mốc để pack lại Tốc độ đúng chỗ (ngay trên Độ mờ) khi hiện lại
        self._logo_speed_anchor = ctk.CTkFrame(pane, height=0,
                                               fg_color="transparent")
        self._logo_speed_anchor.pack(fill="x")

        # Độ mờ — áp dụng cho logo cả khi đứng yên lẫn khi chuyển động
        op_row = ctk.CTkFrame(pane, fg_color="transparent")
        op_row.pack(fill="x", pady=2)
        ctk.CTkLabel(op_row, text="Độ mờ:", width=90,
                     anchor="w").pack(side="left")
        self._logo_opacity = ctk.CTkOptionMenu(
            op_row, values=["Rõ", "Mờ vừa", "Mờ nhiều"], width=165)
        self._logo_opacity.pack(side="right")

        self._logo_size = self._labeled_entry(pane, "Kích thước (%):", "15")

        # Thiết lập độ phân giải
        ctk.CTkLabel(pane, text="Độ phân giải video:", anchor="w").pack(anchor="w", pady=(10, 2))
        self._resolution_var = StringVar(value="Nâng độ phân giải lên 1080x1920")
        self._resolution_menu = ctk.CTkOptionMenu(
            pane,
            values=["Nâng độ phân giải lên 1080x1920", "Giữ nguyên độ phân giải video"],
            variable=self._resolution_var,
            width=230,
        )
        self._resolution_menu.pack(fill="x", pady=(0, 6))

        self._toggle_logo_speed()  # áp dụng trạng thái ẩn ban đầu

        # ── Nơi lưu video: Google Drive hoặc Local ──
        self._section(pane, "  Nơi lưu video")

        self._save_mode = ctk.CTkSegmentedButton(
            pane, values=["Google Drive", "Local"],
            command=self._toggle_save_mode,
        )
        self._save_mode.set("Google Drive")
        self._save_mode.pack(fill="x", pady=(0, 6))

        # Khối Drive (hiện khi chọn "Google Drive")
        self._drive_box = ctk.CTkFrame(pane, fg_color="transparent")
        self._login_label = ctk.CTkLabel(
            self._drive_box, text="Chưa đăng nhập",
            text_color="gray", font=("", 11),
        )
        self._login_label.pack(anchor="w", pady=(0, 6))

        btn_row = ctk.CTkFrame(self._drive_box, fg_color="transparent")
        btn_row.pack(fill="x", pady=(0, 6))
        ctk.CTkButton(btn_row, text=" Đăng nhập", height=34,
                      fg_color="#2d6a4f",
                      command=self._login_drive).pack(side="left", fill="x",
                                                       expand=True, padx=(0, 2))
        ctk.CTkButton(btn_row, text=" Đăng xuất", height=34, width=90,
                      fg_color="#555",
                      command=self._logout_drive).pack(side="right")

        # Khối Local (hiện khi chọn "Local")
        self._local_box = ctk.CTkFrame(pane, fg_color="transparent")
        self._local_label = ctk.CTkLabel(
            self._local_box, text="Chưa chọn thư mục",
            text_color="gray", font=("", 11), wraplength=230, justify="left",
        )
        self._local_label.pack(anchor="w", pady=(0, 4))
        ctk.CTkButton(self._local_box, text=" Chọn thư mục lưu", height=34,
                      fg_color="#2d6a4f",
                      command=self._pick_local_dir).pack(fill="x", pady=(0, 6))

        # Mốc để pack lại khối chế độ đúng chỗ (ngay trên "Tên folder")
        self._save_anchor = ctk.CTkFrame(pane, height=0, fg_color="transparent")
        self._save_anchor.pack(fill="x")

        # Tên folder: dùng chung — Drive = tên folder Drive; Local = tên thư mục con
        folder_row = ctk.CTkFrame(pane, fg_color="transparent")
        folder_row.pack(fill="x", pady=2)
        ctk.CTkLabel(folder_row, text="Tên folder:", anchor="w",
                     width=85).pack(side="left")
        self._folder_name = ctk.CTkEntry(folder_row, width=145,
                                          placeholder_text="VD: Reup_2026")
        self._folder_name.pack(side="right")

        self._toggle_save_mode()  # áp dụng trạng thái hiển thị ban đầu

        # ── Buttons ──
        ctk.CTkFrame(pane, height=1, fg_color="#444").pack(fill="x", pady=16)

        self._btn_process = ctk.CTkButton(
            pane, text="XỬ LÝ VIDEO", height=44,
            font=("", 14, "bold"), fg_color="#1565C0",
            command=self._start_processing,
        )
        self._btn_process.pack(fill="x", pady=(0, 6))

        self._btn_stop = ctk.CTkButton(
            pane, text="DỪNG", height=34,
            fg_color="#8B0000", state="disabled",
            command=self._request_stop,
        )
        self._btn_stop.pack(fill="x")

        # ── Bản quyền ──
        ctk.CTkFrame(pane, height=1, fg_color="#444").pack(fill="x", pady=12)
        data = self._license_data or {}
        key  = data.get("key") or ""
        masked = ("••••" + key[-4:]) if len(key) >= 4 else (key or "N/A")
        expiry = data.get("expire_date") or "Vô thời hạn"
        ctk.CTkLabel(
            pane, text=f" Bản quyền: {masked}\nHết hạn: {expiry}",
            text_color="gray", font=("", 11), justify="left",
        ).pack(anchor="w")
        ctk.CTkButton(
            pane, text="Hủy kích hoạt", height=28, fg_color="#555",
            font=("", 11), command=self._deactivate_license,
        ).pack(anchor="w", pady=(4, 0))

    # ── Video list ───────────────────────────────────────────────────────────

    def _build_video_list(self):
        container = ctk.CTkFrame(self)
        container.grid(row=1, column=1, sticky="nsew", padx=(4, 8), pady=8)
        container.grid_rowconfigure(1, weight=1)
        container.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(container, height=40, corner_radius=0,
                            fg_color="#1a1a2e")
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_propagate(False)

        self._list_title = ctk.CTkLabel(hdr, text="  Danh sách video  (0)",
                                         font=("", 13, "bold"))
        self._list_title.pack(side="left", padx=12)

        ctk.CTkButton(hdr, text="[v] Tất cả", width=80, height=28,
                      font=("", 11),
                      command=lambda: self._select_all(True)).pack(
                          side="right", padx=4)
        ctk.CTkButton(hdr, text="[ ] Bỏ hết", width=80, height=28,
                      font=("", 11), fg_color="#555",
                      command=lambda: self._select_all(False)).pack(
                          side="right", padx=4)

        self._scroll = ctk.CTkScrollableFrame(container, fg_color="#1e1e1e")
        self._scroll.grid(row=1, column=0, sticky="nsew", padx=2, pady=2)
        self._scroll.grid_columnconfigure(0, weight=1)

        self._empty_lbl = ctk.CTkLabel(
            self._scroll,
            text="  Chưa có video\nBấm  'Import CSV'  để bắt đầu",
            font=("", 14), text_color="gray",
        )
        self._empty_lbl.pack(expand=True, pady=80)

        # ── Thanh phân trang ──
        pager = ctk.CTkFrame(container, height=40, corner_radius=0,
                             fg_color="#1a1a2e")
        pager.grid(row=2, column=0, sticky="ew")
        pager.grid_propagate(False)

        self._btn_prev = ctk.CTkButton(pager, text="< Trước", width=84, height=28,
                                       font=("", 11), command=self._prev_page)
        self._btn_prev.pack(side="left", padx=(10, 4), pady=6)

        self._btn_next = ctk.CTkButton(pager, text="Sau >", width=84, height=28,
                                       font=("", 11), command=self._next_page)
        self._btn_next.pack(side="left", padx=4, pady=6)

        self._page_lbl = ctk.CTkLabel(pager, text="", font=("", 11),
                                      text_color="gray")
        self._page_lbl.pack(side="left", padx=12)

        self._update_pager()

    # ── Bottom bar ───────────────────────────────────────────────────────────

    def _build_bottom(self):
        # Log có phân loại: lưu toàn bộ dòng kèm nhãn, dropdown chỉ lọc hiển thị.
        self._log_entries: list[tuple] = []          # (cat | None, text)
        self._log_filter = StringVar(value=LOG_FILTER_ALL)
        self._log_bar_h = 285                         # chiều cao thanh dưới (kéo được)

        bar = ctk.CTkFrame(self, height=self._log_bar_h, corner_radius=0)
        bar.grid(row=2, column=0, columnspan=2, sticky="ew")
        bar.grid_propagate(False)
        bar.grid_columnconfigure(0, weight=1)
        bar.grid_rowconfigure(3, weight=1)            # hàng log co giãn theo bar
        self._bottom_bar = bar

        # (0) Tay nắm kéo: rê lên/xuống để phóng to/thu nhỏ khối log
        grip = ctk.CTkFrame(bar, height=7, corner_radius=0, fg_color="#4a4a4a")
        grip.grid(row=0, column=0, sticky="ew")
        grip.bind("<Enter>", lambda e: grip.configure(fg_color="#6a6a6a"))
        grip.bind("<Leave>", lambda e: grip.configure(fg_color="#4a4a4a"))
        grip.bind("<B1-Motion>", self._on_log_resize)
        try:
            grip.configure(cursor="sb_v_double_arrow")
        except Exception:   # noqa: BLE001
            pass

        # (1) Tiến độ tổng
        prow = ctk.CTkFrame(bar, fg_color="transparent")
        prow.grid(row=1, column=0, sticky="ew", padx=12, pady=(8, 2))
        prow.grid_columnconfigure(0, weight=1)

        self._progress_lbl = ctk.CTkLabel(prow, text="Sẵn sàng",
                                           anchor="w", font=("", 12))
        self._progress_lbl.grid(row=0, column=0, sticky="w")

        self._progress_bar = ctk.CTkProgressBar(prow, height=12)
        self._progress_bar.grid(row=1, column=0, sticky="ew", pady=(2, 0))
        self._progress_bar.set(0)

        self._progress_pct = ctk.CTkLabel(prow, text="0 %", width=55,
                                           font=("", 11))
        self._progress_pct.grid(row=1, column=1, padx=8)

        # (2) Hàng tiêu đề log + dropdown lọc
        hrow = ctk.CTkFrame(bar, fg_color="transparent")
        hrow.grid(row=2, column=0, sticky="ew", padx=12, pady=(2, 0))
        hrow.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(hrow, text="Nhật ký:", font=("", 12)).grid(
            row=0, column=0, sticky="w")
        ctk.CTkOptionMenu(
            hrow, variable=self._log_filter, values=list(LOG_FILTER_OPTS),
            width=140, height=26, font=("", 11),
            command=lambda _v: self._render_log(),
        ).grid(row=0, column=2, sticky="e")

        # (3) Khung log
        # "Consolas" không có trên Ubuntu → Xft fallback to large font → X11 BadLength
        _log_font = ("Consolas", 12) if sys.platform == "win32" else ("", 11)
        self._log = ctk.CTkTextbox(bar, height=180, font=_log_font,
                                    state="disabled")
        self._log.grid(row=3, column=0, sticky="nsew", padx=12, pady=(2, 8))

    # ── Resize khối log ────────────────────────────────────────────────────────

    def _on_log_resize(self, event):
        """Rê tay nắm: đặt lại chiều cao thanh dưới sao cho mép trên bám con trỏ."""
        win_bottom = self.winfo_rooty() + self.winfo_height()
        new_h = win_bottom - event.y_root
        new_h = max(150, min(new_h, self.winfo_height() - 160))
        self._log_bar_h = new_h
        self._bottom_bar.configure(height=new_h)

    # ── Helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _section(parent, text: str):
        ctk.CTkLabel(parent, text=text, font=("", 13, "bold")).pack(
            anchor="w", pady=(14, 4))

    @staticmethod
    def _labeled_entry(parent, label: str, default: str) -> ctk.CTkEntry:
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=2)
        ctk.CTkLabel(row, text=label, anchor="w").pack(side="left")
        e = ctk.CTkEntry(row, width=75)
        e.pack(side="right")
        e.insert(0, default)
        return e

    @staticmethod
    def _fullwidth_entry(parent, label: str, *, show: str = "",
                         placeholder: str = "", default: str = "") -> ctk.CTkEntry:
        """Label ở trên, ô nhập full-width ở dưới (cho API key / model)."""
        ctk.CTkLabel(parent, text=label, anchor="w").pack(anchor="w", pady=(6, 1))
        e = ctk.CTkEntry(parent, placeholder_text=placeholder,
                         show=show or "")
        e.pack(fill="x")
        if default:
            e.insert(0, default)
        return e

    def _toggle_voiceai(self):
        """Hiện/ẩn khối Voice AI và vô hiệu hoá audio cũ khi bật."""
        on = self._voiceai_var.get()
        if on:
            self._voiceai_box.pack(fill="x", before=self._voiceai_anchor)
        else:
            self._voiceai_box.pack_forget()
        # Voice AI ưu tiên tuyệt đối: khoá 'Chọn Audio' / 'Tắt tiếng' (không fallback)
        state = "disabled" if on else "normal"
        for w in (self._btn_pick_audio, self._btn_clear_audio, self._mute_chk):
            try:
                w.configure(state=state)
            except Exception:
                pass

    def _toggle_tts_provider(self, *_):
        """Hiện khối Google TTS, AutoVoice hoặc Voice API theo dropdown nhà cung cấp."""
        self._google_box.pack_forget()
        self._autovoice_box.pack_forget()
        self._videoai_box.pack_forget()
        val = self._tts_provider.get()
        if val == "AutoVoice.vn":
            box = self._autovoice_box
        elif val == "Voice API (videoai)":
            box = self._videoai_box
        else:
            box = self._google_box
        box.pack(fill="x", before=self._tts_provider_anchor)

    def _toggle_ai_provider(self, *_):
        """Hiện khối Gemini hoặc ChatGPT theo dropdown AI tạo text."""
        self._gemini_box.pack_forget()
        self._chatgpt_box.pack_forget()
        box = (self._chatgpt_box
               if self._ai_provider.get() == "ChatGPT"
               else self._gemini_box)
        box.pack(fill="x", before=self._ai_provider_anchor)

    def _toggle_logo_speed(self, *_):
        """Hiện ô Tốc độ chỉ khi logo có chuyển động (khác 'Cố định')."""
        if self._logo_motion.get() == "Cố định":
            self._logo_speed_row.pack_forget()
        else:
            self._logo_speed_row.pack(fill="x", pady=2,
                                      before=self._logo_speed_anchor)

    def _toggle_save_mode(self, *_):
        """Hiện khối Drive hoặc Local theo nút gạt 'Nơi lưu video'."""
        self._drive_box.pack_forget()
        self._local_box.pack_forget()
        box = (self._local_box if self._save_mode.get() == "Local"
               else self._drive_box)
        box.pack(fill="x", before=self._save_anchor)

    def _deactivate_license(self):
        """Xoá key bản quyền khỏi máy rồi thoát (mở lại sẽ phải nhập key)."""
        if not messagebox.askyesno(
                "Hủy kích hoạt",
                "Xoá key bản quyền khỏi máy này và thoát ứng dụng?\n"
                "Lần mở sau sẽ phải nhập key lại."):
            return
        license_mod.clear_key()
        self.destroy()

    def _ui(self, func):
        """Đẩy func về luồng UI. An toàn khi gọi từ background thread kể cả sau khi cửa sổ đóng."""
        try:
            if self.winfo_exists():
                self.after(0, func)
        except Exception:
            pass

    # ── Kiểm tra license định kỳ ──────────────────────────────────────────────

    def _schedule_recheck(self):
        """Chạy trên luồng UI. Bắn 1 thread nền gọi API (không làm đơ UI)."""
        self._recheck_job = None
        threading.Thread(target=self._run_recheck_bg, daemon=True).start()

    def _run_recheck_bg(self):
        """
        Chạy trên LUỒNG NỀN — tuyệt đối không chạm widget Tkinter ở đây
        (kể cả winfo_exists). Chỉ gọi API rồi đẩy kết quả về luồng UI.
        """
        status = license_mod.recheck_license()       # block ~15s, an toàn vì ở nền
        self.after(0, lambda: self._on_recheck_result(status))

    def _on_recheck_result(self, status: str):
        """Chạy trên luồng UI chính — xử lý 3 kịch bản."""
        if not self.winfo_exists():                  # app đã đóng giữa chừng
            return

        if status == license_mod.RECHECK_VALID:
            # KB1: vẫn hợp lệ → reset retry, hẹn chu kỳ 30' tiếp theo
            self._recheck_retries = 0
            self._recheck_job = self.after(_RECHECK_SECONDS * 1000,
                                           self._schedule_recheck)

        elif status == license_mod.RECHECK_NETERR:
            # KB3: lỗi mạng → retry tối đa _MAX_RETRIES lần, mỗi lần cách 1'
            self._recheck_retries += 1
            if self._recheck_retries <= _MAX_RETRIES:
                self._recheck_job = self.after(_RETRY_SECONDS * 1000,
                                               self._schedule_recheck)
            else:
                # Hết lượt retry mà vẫn mất mạng → coi như không có bản quyền
                self._lock_ui_license(
                    "Không thể xác minh bản quyền (mất kết nối mạng).\n"
                    "Vui lòng kiểm tra mạng và mở lại ứng dụng.")

        else:  # RECHECK_INVALID — recheck_license() đã clear_key()
            # KB2: key hết hạn/bị thu hồi
            self._lock_ui_license(
                "Bản quyền của bạn đã hết hạn hoặc bị thu hồi.\n"
                "Vui lòng kích hoạt lại key hợp lệ.")

    def _lock_ui_license(self, message: str):
        """Khóa toàn bộ giao diện bằng overlay phủ kín cửa sổ."""
        if self._license_locked:
            return
        self._license_locked = True

        # Hủy job đang chờ để không recheck tiếp
        if self._recheck_job is not None:
            try:
                self.after_cancel(self._recheck_job)
            except Exception:
                pass
            self._recheck_job = None

        # Dừng tác vụ xử lý video đang chạy (nếu có)
        self._stop_flag = True

        # Overlay che toàn cửa sổ → chặn mọi tương tác bên dưới
        overlay = ctk.CTkFrame(self, fg_color="#0d0d1a")
        overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        ctk.CTkLabel(overlay, text=" Bản quyền không hợp lệ",
                     font=("", 22, "bold"), text_color="#ff4757"
                     ).pack(pady=(180, 10))
        ctk.CTkLabel(overlay, text=message, font=("", 14),
                     wraplength=600, justify="center").pack(pady=10)
        ctk.CTkButton(overlay, text="Kích hoạt lại / Thoát",
                      width=240, height=40,
                      command=self._relaunch_to_license).pack(pady=20)

        # Thông báo lỗi rõ ràng cho người dùng
        messagebox.showerror("Bản quyền", message)

    def _relaunch_to_license(self):
        """Đóng App và yêu cầu main() mở lại LicenseDialog."""
        self._relaunch_license = True
        self.destroy()

    def destroy(self):
        # Hủy job đang chờ trước khi đóng để tránh callback trên widget đã hủy
        if getattr(self, "_recheck_job", None) is not None:
            try:
                self.after_cancel(self._recheck_job)
            except Exception:
                pass
            self._recheck_job = None
        super().destroy()

    def _is_log_match(self, sel: str, cat, text: str) -> bool:
        if sel == LOG_FILTER_ALL:
            return True
        if sel == LOG_ERROR:
            if cat == LOG_ERROR:
                return True
            t = (text or "").strip()
            return t.startswith("X") or t.startswith("!") or "lỗi" in t.lower()
        return sel == cat

    def _is_at_bottom(self) -> bool:
        try:
            if hasattr(self._log, "_textbox"):
                _, y2 = self._log._textbox.yview()
            else:
                _, y2 = self._log.yview()
            return y2 >= 0.95
        except Exception:
            return True

    def _log_add(self, text: str, cat=None):
        """
        Thêm 1 dòng log kèm nhãn phân loại `cat` (None = dòng hệ thống, chỉ hiện
        ở mục "Tất cả"). Chỉ chèn vào textbox khi khớp bộ lọc đang chọn.
        Nếu người dùng đang cuộn lên trên xem log cũ, không kéo xuống dưới.
        """
        def _do():
            self._log_entries.append((cat, text))
            sel = self._log_filter.get()
            if self._is_log_match(sel, cat, text):
                at_bottom = self._is_at_bottom()
                self._log.configure(state="normal")
                # Chunk insert to avoid X11 BadLength crash
                full_text = text + "\n"
                chunk_size = 4000
                for i in range(0, len(full_text), chunk_size):
                    self._log.insert("end", full_text[i:i+chunk_size])
                if at_bottom:
                    self._log.see("end")
                self._log.configure(state="disabled")
        self._ui(_do)

    def _log_msg(self, msg: str, cat=None):
        """Tương thích cũ: tách chuỗi nhiều dòng thành từng dòng log."""
        for line in str(msg).split("\n"):
            self._log_add(line, cat)

    def _render_log(self):
        """Dựng lại nội dung textbox theo bộ lọc hiện tại (khi đổi dropdown)."""
        sel = self._log_filter.get()
        lines = [t for c, t in self._log_entries if self._is_log_match(sel, c, t)]
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        if lines:
            # Chunk insert to avoid X11 BadLength crash on Linux (RenderAddGlyphs)
            full_text = "\n".join(lines) + "\n"
            chunk_size = 4000
            for i in range(0, len(full_text), chunk_size):
                self._log.insert("end", full_text[i:i+chunk_size])
        self._log.see("end")
        self._log.configure(state="disabled")

    # ── File pickers ─────────────────────────────────────────────────────────

    def _import_file(self):
        path = filedialog.askopenfilename(
            title="Chọn file CSV hoặc Excel",
            filetypes=[
                ("CSV/Excel", "*.csv *.xlsx *.xls"),
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx *.xls"),
                ("All", "*.*")
            ],
        )
        if not path:
            return
        try:
            ext = Path(path).suffix.lower()
            if ext in (".xlsx", ".xls"):
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True, data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                     raise ValueError("File Excel trống!")
                headers = [str(h).strip() if h is not None else f"column_{i}" for i, h in enumerate(rows[0])]
                self._videos = []
                for r in rows[1:]:
                     if any(v is not None for v in r):
                         row_dict = {}
                         for h, v in zip(headers, r):
                             row_dict[h] = str(v).strip() if v is not None else ""
                         self._videos.append(row_dict)
            else:
                with open(path, encoding="utf-8-sig") as fh:
                    self._videos = [dict(r) for r in csv.DictReader(fh)]
            self._csv_path = path
            # Khởi tạo state song song (mặc định chọn hết, trạng thái chờ)
            self._selected = [True] * len(self._videos)
            self._status   = ["pending"] * len(self._videos)
            self._page     = 0
            self._csv_label.configure(
                text=f"OK  {Path(path).name}  ({len(self._videos)} videos)",
                text_color="#4CAF50",
            )
            self._populate_list()
            self._log_msg(f"OK Import {len(self._videos)} video từ {Path(path).name}")
        except Exception as exc:
            messagebox.showerror("Lỗi đọc File", str(exc))

    def _pick_csv_out(self):
        path = filedialog.asksaveasfilename(
            title="Lưu CSV output",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="video_stats_output.csv",
        )
        if path:
            self._csv_out_path.set(path)
            short = path if len(path) < 45 else "…" + path[-42:]
            self._csv_out_label.configure(text=f"  {short}",
                                           text_color="#4CAF50")

    def _pick_local_dir(self):
        path = filedialog.askdirectory(title="Chọn thư mục lưu video")
        if path:
            self._local_dir.set(path)
            short = path if len(path) < 40 else "…" + path[-37:]
            self._local_label.configure(text=f"  {short}",
                                        text_color="#4CAF50")

    def _pick_audio(self):
        path = filedialog.askopenfilename(
            title="Chọn file audio",
            filetypes=[("Audio", "*.mp3 *.wav *.aac *.m4a *.ogg"),
                       ("All", "*.*")],
        )
        if path:
            self._audio_path.set(path)
            self._audio_label.configure(
                text=f"  {Path(path).name}", text_color="#4CAF50")
            self._mute_var.set(False)

    def _clear_audio(self):
        self._audio_path.set("")
        self._audio_label.configure(text="Chưa chọn file audio",
                                     text_color="gray")

    def _pick_logo(self):
        path = filedialog.askopenfilename(
            title="Chọn file logo",
            filetypes=[("Image", "*.png *.jpg *.jpeg *.webp"),
                       ("All", "*.*")],
        )
        if path:
            self._logo_path.set(path)
            self._logo_label.configure(
                text=f"  {Path(path).name}", text_color="#4CAF50")

    def _clear_logo(self):
        self._logo_path.set("")
        self._logo_label.configure(text="Chưa chọn file logo",
                                    text_color="gray")

    def _pick_creds(self):
        path = filedialog.askopenfilename(
            title="Chọn file credentials.json",
            filetypes=[("JSON", "*.json"), ("All", "*.*")],
        )
        if path:
            self._creds_path.set(path)
            self._creds_label.configure(
                text=f"  {Path(path).name}", text_color="#4CAF50")

    # ── Drive login ──────────────────────────────────────────────────────────

    def _login_drive(self):
        # Cancel any in-progress login, then start fresh
        if self._login_in_progress:
            self._login_cancel_flag.set()

        self._login_cancel_flag = threading.Event()
        cancel_flag = self._login_cancel_flag
        self._login_in_progress = True

        def _do():
            try:
                from drive import DriveUploader
                email = DriveUploader().test_connection()
                if cancel_flag.is_set():
                    return
                self._ui(lambda e=email: [
                    self._login_label.configure(text=f"OK {e}", text_color="#4CAF50"),
                    messagebox.showinfo("Đăng nhập thành công OK", f"Đã đăng nhập:\n{e}"),
                ])
                self._log_msg(f"OK Drive đăng nhập OK — {email}")
            except Exception as exc:
                if cancel_flag.is_set():
                    return
                self._ui(lambda e=exc: [
                    self._login_label.configure(
                        text="Đăng nhập thất bại", text_color="#f44336"),
                    messagebox.showerror("Lỗi đăng nhập X", str(e)),
                ])
            finally:
                if not cancel_flag.is_set():
                    self._login_in_progress = False

        threading.Thread(target=_do, daemon=True).start()

    def _logout_drive(self):
        from drive import get_token_path
        _TOKEN_PATH = get_token_path()
        if _TOKEN_PATH.exists():
            _TOKEN_PATH.unlink()
            self._login_label.configure(text="Đã đăng xuất", text_color="gray")
            self._log_msg(" Đã xoá token — đăng nhập lại để dùng Drive.")
        else:
            self._login_label.configure(text="Chưa đăng nhập", text_color="gray")

    # ── Video list ───────────────────────────────────────────────────────────

    def _page_count(self) -> int:
        total = len(self._videos)
        return max(1, (total + _PAGE_SIZE - 1) // _PAGE_SIZE)

    def _populate_list(self):
        # Chỉ dựng widget cho TRANG hiện tại → tránh treo khi CSV có nghìn video
        for w in self._scroll.winfo_children():
            w.destroy()
        self._rows.clear()
        self._row_by_index.clear()

        if not self._videos:
            self._empty_lbl = ctk.CTkLabel(
                self._scroll,
                text="  Chưa có video\nBấm  'Import CSV'  để bắt đầu",
                font=("", 14), text_color="gray",
            )
            self._empty_lbl.pack(expand=True, pady=80)
            self._update_pager()
            return

        total = len(self._videos)
        self._page = max(0, min(self._page, self._page_count() - 1))
        start = self._page * _PAGE_SIZE
        end   = min(start + _PAGE_SIZE, total)

        self._list_title.configure(text=f"  Danh sách video  ({total})")

        for gidx in range(start, end):
            row = VideoRow(self._scroll, self, self._videos[gidx], gidx)
            row.pack(fill="x", padx=2, pady=1)
            self._rows.append(row)
            self._row_by_index[gidx] = row

        self._update_pager()

    def _update_pager(self):
        total = len(self._videos)
        if total == 0:
            self._page_lbl.configure(text="")
            self._btn_prev.configure(state="disabled")
            self._btn_next.configure(state="disabled")
            return
        pages = self._page_count()
        start = self._page * _PAGE_SIZE + 1
        end   = min((self._page + 1) * _PAGE_SIZE, total)
        self._page_lbl.configure(
            text=f"Trang {self._page + 1}/{pages}   "
                 f"(video {start}–{end} / {total})")
        self._btn_prev.configure(state="normal" if self._page > 0 else "disabled")
        self._btn_next.configure(
            state="normal" if self._page < pages - 1 else "disabled")

    def _prev_page(self):
        if self._page > 0:
            self._page -= 1
            self._populate_list()

    def _next_page(self):
        if self._page < self._page_count() - 1:
            self._page += 1
            self._populate_list()

    def _select_all(self, state: bool):
        # Áp cho TẤT CẢ video (mọi trang), không chỉ trang đang hiển thị
        for i in range(len(self._selected)):
            self._selected[i] = state
        for row in self._rows:           # cập nhật checkbox đang hiển thị
            row.selected.set(state)

    # ── Cập nhật trạng thái/tiến độ 1 video (an toàn khi dòng không hiển thị) ──

    def _set_row_status(self, gidx: int, status: str):
        if 0 <= gidx < len(self._status):
            self._status[gidx] = status        # lưu để giữ qua trang
        row = self._row_by_index.get(gidx)
        if row is not None:
            row.set_status(status)

    def _set_row_progress(self, gidx: int, value: float):
        row = self._row_by_index.get(gidx)
        if row is not None:
            row.set_progress(value)

    # ── Processing ───────────────────────────────────────────────────────────

    def _start_processing(self):
        if not self._videos:
            messagebox.showwarning("Chưa có video", "Import file CSV trước!")
            return
        save_mode = "local" if self._save_mode.get() == "Local" else "drive"
        if not self._folder_name.get().strip():
            messagebox.showwarning(
                "Thiếu tên folder",
                "Nhập tên thư mục lưu video!" if save_mode == "local"
                else "Nhập tên folder trên Google Drive!")
            return
        if save_mode == "local" and not self._local_dir.get():
            messagebox.showwarning("Chưa chọn thư mục",
                                   "Bấm 'Chọn thư mục lưu' để chọn nơi lưu video!")
            return
        if not self._csv_out_path.get():
            messagebox.showwarning("Chưa chọn nơi lưu CSV",
                                   "Bấm 'Lưu CSV' để chọn nơi lưu file output!")
            return

        # Lấy lựa chọn từ model (toàn bộ video, không chỉ trang đang hiển thị)
        selected = [i for i, sel in enumerate(self._selected) if sel]
        if not selected:
            messagebox.showwarning("Chưa chọn", "Tick chọn ít nhất 1 video!")
            return

        try:
            trim_start = float(self._trim_start.get() or "0")
            trim_end   = float(self._trim_end.get() or "0")
            logo_size  = int(self._logo_size.get() or "15")
            if IS_PRO and self._workers and self._delay:
                workers = max(1, min(8, int(self._workers.get() or "2")))
                delay   = max(0.0, float(self._delay.get() or "1"))
            else:
                workers = 1
                delay   = 60.0
        except ValueError:
            messagebox.showerror("Lỗi", "Giây cắt và % logo phải là số!")
            return

        # Voice AI có độ ưu tiên cao nhất; audio sẽ được sinh riêng cho từng video
        voice_ai = None
        if self._voiceai_var.get():
            tts_provider_choice = self._tts_provider.get()
            ai_provider = self._ai_provider.get()
            if ai_provider == "ChatGPT":
                ai_key = self._chatgpt_key.get().strip()
                ai_model = self._chatgpt_model.get().strip() or "gpt-4o-mini"
                key_warn_prefix = "API Key OpenAI"
            else:
                ai_key = self._gemini_key.get().strip()
                ai_model = self._gemini_model.get().strip() or "gemini-3.1-flash-lite"
                key_warn_prefix = "API Key Gemini"

            prompt     = self._prompt_box.get("1.0", "end").strip()

            if tts_provider_choice == "AutoVoice.vn":
                tts_key    = self._autovoice_key.get().strip()
                voice_name = (self._autovoice_voice.get().strip() or DEFAULT_AUTOVOICE_VOICE)
                speed_str  = self._autovoice_speed.get() or "1"
                key_warn   = f"Voice AI cần cả {key_warn_prefix} và X-API-Key AutoVoice!"
                provider   = "autovoice"
            elif tts_provider_choice == "Voice API (videoai)":
                tts_key    = self._videoai_key.get().strip()
                voice_name = (self._videoai_voice.get().strip() or DEFAULT_VIDEOAI_VOICE)
                speed_str  = self._videoai_speed.get() or "1"
                key_warn   = f"Voice AI cần cả {key_warn_prefix} và X-API-Key Voice API!"
                provider   = "videoai"
            else:
                tts_key    = self._tts_key.get().strip()
                voice_name = VOICE_CHOICES.get(
                    self._voice_name.get(), "vi-VN-Standard-C")
                speed_str  = self._tts_speed.get() or "1.2"
                key_warn   = f"Voice AI cần cả {key_warn_prefix} và API Key Google TTS!"
                provider   = "google"

            if not ai_key or not tts_key:
                messagebox.showwarning("Thiếu API key", key_warn)
                return
            if not prompt:
                messagebox.showwarning(
                    "Thiếu prompt", "Nhập prompt để tạo lời thoại!")
                return
            try:
                speed = float(speed_str)
            except ValueError:
                messagebox.showerror("Lỗi", "Tốc độ đọc phải là số!")
                return
            voice_ai = {
                "ai_provider": ai_provider,
                "ai_key":      ai_key,
                "model":       ai_model,
                "provider":    provider,
                "tts_key":     tts_key,
                "voice_name":  voice_name,
                "speed":       speed,
                "prompt":      prompt,
            }

        if voice_ai:
            audio_path = None       # sẽ được gán per-row trong worker
        elif self._mute_var.get():
            audio_path = ""
        elif self._audio_path.get():
            audio_path = self._audio_path.get()
        else:
            audio_path = None

        settings = {
            "trim_start":    trim_start,
            "trim_end":      trim_end,
            "audio_path":    audio_path,
            "voice_ai":      voice_ai,
            "logo_path":     self._logo_path.get() or None,
            "logo_position": self._logo_pos.get(),
            "logo_size":     logo_size,
            "logo_motion":   {"Cố định": "static",
                              "Chạy vòng theo viền": "perimeter",
                              "Nảy DVD": "bounce"}.get(
                                  self._logo_motion.get(), "static"),
            "logo_speed":    {"Chậm": "slow", "Vừa": "normal",
                              "Nhanh": "fast"}.get(
                                  self._logo_speed.get(), "normal"),
            "logo_opacity":  {"Rõ": "opaque", "Mờ vừa": "medium",
                              "Mờ nhiều": "light"}.get(
                                  self._logo_opacity.get(), "opaque"),
            "force_1080p":   self._resolution_var.get() == "Nâng độ phân giải lên 1080x1920",
            "folder_name":   self._folder_name.get().strip(),
            "csv_out":       self._csv_out_path.get(),
            "workers":       workers,
            "delay":         delay,
            "save_mode":     save_mode,
            "local_dir":     self._local_dir.get(),
        }

        self._processing = True
        self._stop_flag  = False
        self._btn_process.configure(state="disabled")
        self._btn_stop.configure(state="normal")
        self._progress_bar.set(0)
        self._progress_pct.configure(text="0 %")

        # Xoá log của lần chạy trước (tránh tích luỹ qua nhiều lần xử lý)
        self._log_entries.clear()
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

        threading.Thread(
            target=self._run_batch,
            args=(selected, settings),
            daemon=True,
        ).start()

    def _request_stop(self):
        self._stop_flag = True
        self._log_msg("[Dung] Dừng sau khi xong video hiện tại…")
        self._btn_stop.configure(state="disabled")

    # ── Batch runner ─────────────────────────────────────────────────────────

    def _run_batch(self, indices: list[int], settings: dict):
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from drive import DriveUploader

        save_mode   = settings.get("save_mode", "drive")
        total       = len(indices)
        out_refs: dict[int, str] = {}      # id(video_dict) → Drive link / local path
        lock        = threading.Lock()
        done        = 0
        errors      = 0
        completed   = 0          # finished workers (done + error)

        workers     = max(1, int(settings.get("workers", 1)))
        delay       = max(0.0, float(settings.get("delay", 2)))
        # Giãn cách KHỞI ĐỘNG: 2 video liên tiếp cách nhau >= `delay` giây khi
        # bắt đầu (tránh dồn request). next_start = mốc monotonic sớm nhất cho
        # video kế; bảo vệ bằng start_lock riêng để không tranh chấp với `lock`.
        start_lock  = threading.Lock()
        next_start  = [0.0]

        self._ui(lambda: self._progress_lbl.configure(
            text=f"Đang xử lý 0 / {total} …"))

        def _abort(msg: str):
            self._log_msg(msg)
            self._ui(lambda: [
                self._btn_process.configure(state="normal"),
                self._btn_stop.configure(state="disabled"),
            ])
            self._processing = False

        # ── Chuẩn bị đích lưu (1 lần, trước khi spawn worker) ──
        folder_id = None                 # Drive
        dest_dir  = None                 # Local
        reserved: set[str] = set()       # tên file đã đặt chỗ (Local, tránh đua worker)

        if save_mode == "local":
            dest_dir = os.path.join(settings["local_dir"],
                                    settings["folder_name"])
            try:
                os.makedirs(dest_dir, exist_ok=True)
            except Exception as exc:
                _abort(f"X Không tạo được thư mục lưu: {exc}")
                return
            self._log_msg(f" Lưu video vào: {dest_dir}")
        else:
            try:
                self._log_msg("@ Đang kết nối Google Drive…")
                init_uploader = DriveUploader()
                folder_id = init_uploader.create_folder(settings["folder_name"])
                self._log_msg(
                    f"OK Đã tạo folder '{settings['folder_name']}' trên Drive")
            except Exception as exc:
                _abort(f"X Không thể kết nối Drive: {exc}")
                return

        tmp_dir = tempfile.mkdtemp(prefix="reup_")

        # Semaphore giới hạn số FFmpeg process chạy đồng thời.
        # Dù workers = 8, tối đa 4 FFmpeg chạy song song để tránh I/O
        # bão hòa + Windows handle limit (giai đoạn download vẫn song song).
        _ffmpeg_sem = threading.Semaphore(min(workers, 4))

        # Cổng giãn cách: đặt chỗ một "khe khởi động" cách video trước >= delay
        # giây, rồi chờ tới khe đó. Atomic dưới start_lock nên dù nhiều luồng gọi
        # cùng lúc, các khe vẫn xếp so le delay giây.
        def _gate_start():
            if delay <= 0:
                return
            with start_lock:
                slot = max(time.monotonic(), next_start[0])
                next_start[0] = slot + delay
            while not self._stop_flag:
                remaining = slot - time.monotonic()
                if remaining <= 0:
                    break
                time.sleep(min(0.2, remaining))

        # ── Per-video worker ──────────────────────────────────────────────────
        # idx  = thứ tự trong batch (đặt tên temp + log + tiến độ tổng)
        # gidx = chỉ số video trong self._videos (định tuyến trạng thái/lưu CSV)
        def process_one(idx: int, gidx: int):
            nonlocal done, errors, completed

            if self._stop_flag:
                return

            vid     = self._videos[gidx]
            item_id = (vid.get("item_id") or "").strip() or f"video_{idx}"
            url     = (vid.get("video_url") or "").strip()
            val_p   = vid.get("product_name") if vid.get("product_name") is not None else vid.get("productName")
            p_name  = "" if val_p is None or str(val_p).strip().lower() in ("", "nan", "none") else str(val_p).strip()
            name    = (p_name if p_name else "(dữ liệu product_name rỗng)")[:50]
            # idx is unique within this batch → no temp collision on dup item_id
            tmp_dl  = os.path.join(tmp_dir, f"dl_{idx}.mp4")
            tmp_out = os.path.join(tmp_dir, f"out_{idx}.mp4")

            if not url:
                with lock:
                    errors += 1
                    completed += 1
                    _d, _e, _c = done, errors, completed
                    out_refs[id(vid)] = "Lỗi"
                self._ui(lambda g=gidx: self._set_row_status(g, "error"))
                self._log_msg(f"X  [{idx+1}/{total}] Bỏ qua (thiếu video_url): {name}",
                              LOG_ERROR)
                pct = _c / total
                self._ui(lambda p=pct, d=_d, e=_e, c=_c: [
                    self._progress_bar.set(p),
                    self._progress_pct.configure(text=f"{int(p*100)} %"),
                    self._progress_lbl.configure(
                        text=f"Xong {c}/{total}  OK {d}  X {e}"),
                ])
                return

            # Each worker owns its Drive client (not thread-safe to share);
            # chế độ Local không cần Drive.
            worker_uploader = DriveUploader() if save_mode == "drive" else None

            try:
                if self._stop_flag:
                    return
                # Giãn cách khởi động (hàng đợi liên tục): chờ tới khe của video
                # này để không nổ request cùng lúc với video khác.
                _gate_start()
                if self._stop_flag:
                    return
                # 1. Download
                def _dl_log(m):
                    prefix = f"[{idx+1}/{total}] "
                    msg = m if m.startswith(prefix) else f"{prefix}{m}"
                    cat = LOG_ERROR if ("lỗi" in m.lower() or "fail" in m.lower() or "không" in m.lower()) else LOG_DOWNLOAD
                    self._log_msg(msg, cat)

                self._ui(lambda g=gidx: self._set_row_status(g, "downloading"))
                self._downloader.download(
                    url, tmp_dl,
                    progress_cb=lambda p, g=gidx: self._ui(
                        lambda: self._set_row_progress(g, p)),
                    log=_dl_log,
                    should_stop=lambda: self._stop_flag,
                )
                self._log_add(f"OK  [{idx+1}/{total}] Tải xong: {name}", LOG_DOWNLOAD)
                if self._stop_flag:
                    return

                # 2. Voice AI (tuỳ chọn): prompt + dữ liệu CSV → Gemini → TTS
                #    Nếu lỗi sau toàn bộ retry → raise để bỏ qua video hoàn toàn.
                #    Không fallback về audio gốc.
                row_settings = settings
                voice_ai = settings.get("voice_ai")
                voice_mp3 = None
                if voice_ai:
                    self._ui(lambda g=gidx: self._set_row_status(g, "processing"))
                    out_mp3 = os.path.join(tmp_dir, f"voice_{idx}.mp3")
                    row_settings = dict(settings)   # copy riêng, không sửa settings chung

                    prompt_txt = voice_ai.get("prompt", "")
                    val = vid.get("product_name") if vid.get("product_name") is not None else vid.get("productName")
                    p_name = "" if val is None or str(val).strip().lower() in ("", "nan", "none") else str(val).strip()

                    # Lời thoại sinh xong → log nhãn "Tạo text"; retry Gemini cũng
                    # về "Tạo text", còn lại (TTS) về "Tạo voice".
                    def _on_script(s, i=idx):
                        snip = " ".join(str(s).split())[:60]
                        self._log_add(
                            f'OK  [{i+1}/{total}] Tạo text xong: "{snip}…"', LOG_TEXT)

                    def _voice_log(m):
                        prefix = f"[{idx+1}/{total}] "
                        msg = m if m.startswith(prefix) else f"{prefix}{m}"
                        is_err = "lỗi" in m.lower() or "fail" in m.lower() or "too large" in m.lower()
                        cat = LOG_ERROR if is_err else (LOG_TEXT if ("Gemini" in m or "ChatGPT" in m) else LOG_VOICE)
                        self._log_msg(msg, cat)

                    if re.search(r"\$?\{\s*(product_name|productName)\s*\}", prompt_txt) and not p_name:
                        raise RuntimeError(
                            f"Voice AI bỏ qua: ô 'product_name' bị rỗng trong CSV")

                    make_voice(
                        vid, out_mp3,
                        ai_provider=voice_ai.get("ai_provider", "Gemini"),
                        ai_key=voice_ai["ai_key"],
                        tts_key=voice_ai["tts_key"],
                        prompt=prompt_txt,
                        model=voice_ai["model"],
                        provider=voice_ai.get("provider", "google"),
                        voice_name=voice_ai["voice_name"],
                        speaking_rate=voice_ai["speed"],
                        retries=10,
                        tts_retries=20,
                        tts_backoff=2,
                        log=_voice_log,
                        on_script=_on_script,
                        should_stop=lambda: self._stop_flag,
                    )
                    voice_mp3 = out_mp3
                    row_settings["audio_path"] = out_mp3   # thay audio gốc
                    self._log_add(
                        f"OK  [{idx+1}/{total}] Tạo voice xong: {name}", LOG_VOICE)

                if self._stop_flag:
                    if voice_mp3:
                        _safe_remove(voice_mp3)
                    return

                # 3. FFmpeg
                self._ui(lambda g=gidx: self._set_row_status(g, "processing"))
                # Giới hạn số FFmpeg song song (semaphore): download vẫn chạy
                # tối đa `workers` luồng, nhưng FFmpeg encode tối đa 4 cài cùng lúc.
                with _ffmpeg_sem:
                    self._processor.process_video(tmp_dl, tmp_out, row_settings)
                self._log_add(f"OK  [{idx+1}/{total}] Xử lý xong: {name}", LOG_PROCESS)
                _safe_remove(tmp_dl)
                if voice_mp3:
                    _safe_remove(voice_mp3)
                if self._stop_flag:
                    _safe_remove(tmp_out)
                    return

                # 4. Lưu kết quả: upload Drive HOẶC move vào thư mục local
                self._ui(lambda g=gidx: self._set_row_status(g, "uploading"))
                if save_mode == "local":
                    # Đặt tên duy nhất trong lock để worker không giành tên nhau
                    with lock:
                        final_path = _reserve_local_path(
                            dest_dir, f"{item_id}.mp4", reserved)
                    shutil.move(tmp_out, final_path)
                    ref = os.path.abspath(final_path)
                    self._log_add(f"OK  [{idx+1}/{total}] Lưu xong: {name}", LOG_UPLOAD)
                else:
                    ref = worker_uploader.upload_video(
                        tmp_out, filename=f"{item_id}.mp4", folder_id=folder_id,
                        progress_cb=lambda p, g=gidx: self._ui(
                            lambda: self._set_row_progress(g, p)),
                    )
                    _safe_remove(tmp_out)
                    self._log_add(
                        f"OK  [{idx+1}/{total}] Upload xong: {name}", LOG_UPLOAD)

                with lock:
                    # Key by dict identity → robust against duplicate item_id
                    out_refs[id(vid)] = ref
                    done += 1
                self._ui(lambda g=gidx: self._set_row_status(g, "done"))
                self._log_msg(f"OK  [{idx+1}/{total}] Xong: {name}", LOG_RESULT)

            except Exception as exc:
                _safe_remove(tmp_dl)
                _safe_remove(tmp_out)
                if voice_mp3:
                    _safe_remove(voice_mp3)
                with lock:
                    errors += 1
                    out_refs[id(vid)] = "Lỗi"   # ghi vào CSV output cột Link Video
                self._ui(lambda g=gidx: self._set_row_status(g, "error"))
                exc_summary = " ".join([line.strip() for line in str(exc).splitlines() if line.strip()])
                self._log_msg(f"X  [{idx+1}/{total}] Lỗi: {name} → {exc_summary}",
                              LOG_ERROR)

            finally:
                with lock:
                    completed += 1
                    _d, _e, _c = done, errors, completed
                pct = _c / total
                self._ui(lambda p=pct, d=_d, e=_e, c=_c: [
                    self._progress_bar.set(p),
                    self._progress_pct.configure(text=f"{int(p*100)} %"),
                    self._progress_lbl.configure(
                        text=f"Xong {c}/{total}  OK {d}  X {e}"
                    ),
                ])

        # ── Run workers ───────────────────────────────────────────────────────
        # Bọc trong try/finally để LUÔN bật lại nút (kể cả khi Dừng hoặc lỗi),
        # nếu không nút 'XỬ LÝ VIDEO' sẽ kẹt disabled và không chạy lại được.
        try:
            # Hàng đợi liên tục: nạp TẤT CẢ video vào 1 pool `workers` luồng;
            # luồng nào xong là bốc video kế tiếp ngay (không chờ theo đợt nên
            # luôn đủ `workers` luồng chạy). `delay` được áp trong _gate_start()
            # như khoảng giãn cách KHỞI ĐỘNG giữa các video → trải đều request,
            # không dồn cùng lúc.
            with ThreadPoolExecutor(max_workers=workers) as executor:
                futures = [
                    executor.submit(process_one, b, gidx)
                    for b, gidx in enumerate(indices)
                ]
                for _ in as_completed(futures):
                    pass   # progress is updated inside process_one

            # ── Cleanup ───────────────────────────────────────────────────────
            _safe_rmdir(tmp_dir)

            if out_refs:
                try:
                    self._write_output_csv(settings["csv_out"], out_refs)
                    self._log_msg(f"\n Đã lưu CSV: {settings['csv_out']}")
                except Exception as exc:
                    self._log_msg(f"X Lỗi lưu CSV: {exc}")

            if self._stop_flag:
                self._log_msg(
                    f"\n[Dung] Đã dừng.  {done}/{total} hoàn tất trước khi dừng"
                    + (f"  X {errors} lỗi" if errors else "")
                )
            else:
                self._log_msg(
                    f"\n Hoàn thành!  {done}/{total} thành công"
                    + (f"  X {errors} lỗi" if errors else "")
                )
                if done:
                    if save_mode == "local":
                        self._ui(lambda d=done, t=total, dd=dest_dir,
                                 cv=settings["csv_out"]:
                                 self._finish_local(d, t, dd, cv))
                    else:
                        self._ui(lambda: messagebox.showinfo(
                            "Hoàn thành!",
                            f"OK  {done}/{total} video thành công\n"
                            f"@  Folder: {settings['folder_name']}\n"
                            f"  CSV: {settings['csv_out']}",
                        ))
        finally:
            self._ui(lambda: [
                self._btn_process.configure(state="normal"),
                self._btn_stop.configure(state="disabled"),
            ])
            self._stop_flag  = False
            self._processing = False

    def _finish_local(self, done: int, total: int, dest_dir: str, csv_out: str):
        """Thông báo hoàn thành (Local) và hỏi mở thư mục chứa video."""
        open_it = messagebox.askyesno(
            "Hoàn thành!",
            f"OK  {done}/{total} video thành công\n"
            f"  Thư mục: {dest_dir}\n"
            f"  CSV: {csv_out}\n\n"
            "Mở thư mục chứa video?")
        if open_it:
            try:
                if sys.platform == "win32":
                    os.startfile(dest_dir)              # Windows: mở Explorer
                elif sys.platform == "darwin":
                    subprocess.run(["open", dest_dir])  # macOS: Finder
                else:
                    subprocess.run(["xdg-open", dest_dir])  # Linux: file manager
            except Exception as exc:     # noqa: BLE001
                self._log_msg(f"! Không mở được thư mục: {exc}")

    # ── CSV writer ────────────────────────────────────────────────────────────

    def _write_output_csv(self, out_path: str, out_refs: dict):
        """
        Write a new CSV based on input CSV, replacing 'video_url' column with 'Link Video'
        and adding an empty 'hagtag' column to the right of 'Link Video'.
        - Processed video: Drive link OR local path
        - Error: "Lỗi"
        - Unprocessed/stopped: "" (empty)
        """
        if not self._videos:
            return

        # Tạo danh sách fieldnames thay cột video_url bằng "Link Video" và thêm "hagtag" ở bên phải
        fieldnames = []
        for fn in self._videos[0].keys():
            if fn == "video_url":
                if "Link Video" not in fieldnames:
                    fieldnames.append("Link Video")
                    fieldnames.append("hagtag")
            elif fn not in ("Link Video", "hagtag"):
                fieldnames.append(fn)
        if "Link Video" not in fieldnames:
            fieldnames.append("Link Video")
        if "hagtag" not in fieldnames:
            idx = fieldnames.index("Link Video")
            fieldnames.insert(idx + 1, "hagtag")

        with open(out_path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            for vid in self._videos:
                row = dict(vid)
                row.pop("video_url", None)   # Xoá cột video_url
                if id(vid) in out_refs:
                    row["Link Video"] = out_refs[id(vid)]   # link Drive/local path hoặc "Lỗi"
                else:
                    row["Link Video"] = ""                  # Chưa chạy đến thì để trống
                row["hagtag"] = ""                          # Cột hagtag rỗng ở bên phải Link Video
                writer.writerow(row)


# ── Utilities ─────────────────────────────────────────────────────────────────

def _safe_remove(path: str):
    try:
        if path and os.path.exists(path):
            os.remove(path)
    except OSError:
        pass


def _safe_rmdir(path: str):
    try:
        if path and os.path.isdir(path):
            shutil.rmtree(path, ignore_errors=True)
    except OSError:
        pass


def _reserve_local_path(dest_dir: str, filename: str, reserved: set) -> str:
    """
    Trả về đường dẫn trong dest_dir chưa tồn tại và chưa bị đặt chỗ, thêm hậu
    tố _1, _2 … trước phần mở rộng khi trùng. Caller phải giữ lock để các worker
    không chọn trùng tên.
    """
    base, ext = os.path.splitext(filename)
    candidate = os.path.join(dest_dir, filename)
    i = 1
    while candidate in reserved or os.path.exists(candidate):
        candidate = os.path.join(dest_dir, f"{base}_{i}{ext}")
        i += 1
    reserved.add(candidate)
    return candidate


# ── License gate ────────────────────────────────────────────────────────────

class LicenseDialog(ctk.CTk):
    """Cửa sổ nhập key bản quyền — hiện khi chưa kích hoạt hợp lệ."""

    def __init__(self):
        super().__init__()
        _apply_linux_x11_fix(self)
        if sys.platform != "win32":
            try:
                self._check_dpi_scaling = lambda: None
            except Exception:
                pass
        self.activated = False
        self.activated_data = None

        self.title("Kích hoạt bản quyền")
        self.geometry("420x300")
        self.resizable(False, False)

        ctk.CTkLabel(self, text=" Kích hoạt bản quyền",
                     font=("", 18, "bold")).pack(pady=(22, 2))
        ctk.CTkLabel(self, text=APP_TITLE,
                     text_color="gray", font=("", 12)).pack(pady=(0, 14))

        self._entry = ctk.CTkEntry(self, width=340, height=38,
                                   placeholder_text="Nhập key bản quyền…")
        self._entry.pack(pady=(0, 8))
        self._entry.bind("<Return>", lambda _e: self._activate())

        self._btn = ctk.CTkButton(self, text="Kích hoạt", width=340, height=38,
                                  command=self._activate)
        self._btn.pack(pady=(0, 6))

        self._msg = ctk.CTkLabel(self, text="", text_color="#ff4757",
                                 font=("", 12), wraplength=360, justify="left")
        self._msg.pack(pady=(0, 6))

        dev = license_mod.get_device_id()
        ctk.CTkLabel(self, text=f"Mã thiết bị: {dev}",
                     text_color="#666", font=("", 10),
                     wraplength=380, justify="center").pack(side="bottom",
                                                            pady=8)

    def _activate(self):
        key = self._entry.get().strip()
        if not key:
            self._msg.configure(text="Vui lòng nhập key!")
            return
        self._btn.configure(state="disabled", text="Đang kiểm tra…")
        self._msg.configure(text="")

        def work():
            result = license_mod.activate(key)
            self.after(0, lambda: self._done(result))

        threading.Thread(target=work, daemon=True).start()

    def _done(self, result: dict):
        if result.get("is_valid"):
            self.activated = True
            self.activated_data = result.get("data")
            self.destroy()
        else:
            self._msg.configure(
                text=result.get("message") or "Kích hoạt thất bại!")
            self._btn.configure(state="normal", text="Kích hoạt")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    # Bản KHÁCH: chỉ cho mở 1 instance. Bản PRO: cho mở nhiều instance/tab thoải mái
    if not IS_PRO and not single_instance.acquire():
        try:
            root = ctk.CTk()
            _apply_linux_x11_fix(root)
            root.withdraw()
            messagebox.showwarning(
                "Ứng dụng đang chạy",
                f"Ứng dụng {APP_TITLE} đang được mở!\n"
                "Không thể mở nhiều cửa sổ/tab ứng dụng cùng lúc."
            )
            root.destroy()
        except Exception:
            pass
        return

    try:
        # Cổng bản quyền: kiểm tra key lúc mở app (chặt — mất mạng = không vào được)
        while True:
            status = license_mod.check_license()
            if status["is_valid"]:
                app = App(license_data=status["data"])
                app.mainloop()
                # App đóng vì bản quyền bị thu hồi lúc đang chạy → quay lại nhập key
                if getattr(app, "_relaunch_license", False):
                    continue
                return                       # người dùng đóng app bình thường → thoát
            else:
                dlg = LicenseDialog()
                dlg.mainloop()
                if not dlg.activated:        # đóng cửa sổ mà chưa kích hoạt → thoát
                    return
                # Đã kích hoạt → lặp lại, check_license() lần sau sẽ pass
    finally:
        if not IS_PRO:
            single_instance.release()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        tb = traceback.format_exc()
        # Ghi log ra thư mục ghi được (cạnh .AppImage / exe, hoặc ~/.config) để
        # chẩn đoán khi chạy không có console (double-click không hiện traceback).
        _write_crash(tb)
        # In ra stderr (hiện khi chạy từ terminal)
        sys.stderr.write(tb)
        # Cố hiển thị hộp thoại nếu Tk còn dùng được
        try:
            messagebox.showerror("Lỗi khởi động", tb)
        except Exception:
            pass
        raise
